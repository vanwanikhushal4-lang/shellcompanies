#!/usr/bin/env python3
"""Prepare compact JSON inputs for the bootstrap workbook."""

from __future__ import annotations

import hashlib
import csv
import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path("/Users/apple/Downloads/Shell companies")
TEMP = Path("/tmp/company-intelligence.XDfAid")


def text(value):
    return None if value is None else str(value).strip()


queue_book = load_workbook(ROOT / "Com_name&CIN.xlsx", read_only=True, data_only=True)
queue_sheet = queue_book[queue_book.sheetnames[0]]
companies = []
seen = set()
for cin_value, name_value, *_ in queue_sheet.iter_rows(min_row=2, values_only=True):
    cin = text(cin_value)
    name = text(name_value)
    if not cin and not name:
        continue
    cin = (cin or "").upper()
    if cin in seen:
        continue
    seen.add(cin)
    companies.append({"cin": cin, "company_name": name or "", "normalized_name": (name or "").upper()})
queue_book.close()

library = load_workbook(ROOT / "Shell_Company_Indicator_Library_3.xlsx", read_only=True, data_only=True)
fields_sheet = library["Field Dictionary"]
fields = []
for row in fields_sheet.iter_rows(min_row=4, values_only=True):
    if not row[0]:
        continue
    fields.append({
        "field_name": text(row[0]),
        "rules_unlocked": row[1],
        "categories_served": text(row[2]),
        "source_documents": text(row[3]),
        "example_rule_ids": text(row[4]),
    })

documents_sheet = library["Document Catalogue"]
document_priorities = []
for row in documents_sheet.iter_rows(min_row=4, values_only=True):
    if not row[0]:
        continue
    document_priorities.append({
        "source_document": text(row[0]),
        "group": text(row[1]),
        "primary_rules_unlocked": row[2],
        "fallback_uses": row[3],
        "critical": row[4],
        "high": row[5],
        "medium": row[6],
        "low": row[7],
        "categories_served": text(row[8]),
    })
library.close()

local_documents = []
for path in sorted((ROOT / "doc").rglob("*")):
    if not path.is_file() or path.name == ".DS_Store":
        continue
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    relative = path.relative_to(ROOT)
    local_documents.append({
        "company_folder": relative.parts[1] if len(relative.parts) > 1 else "",
        "filename": path.name,
        "extension": path.suffix.lower(),
        "file_size_bytes": path.stat().st_size,
        "sha256": digest.hexdigest(),
        "local_path": str(path),
        "modified_at": datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).isoformat(),
    })

with (ROOT / "intelligence_pipeline/config/free_sources.csv").open(encoding="utf-8-sig", newline="") as handle:
    sources = list(csv.DictReader(handle))

TEMP.mkdir(parents=True, exist_ok=True)
(TEMP / "bootstrap_inputs.json").write_text(json.dumps({
    "generated_at": datetime.now(timezone.utc).isoformat(),
    "companies": companies,
    "fields": fields,
    "document_priorities": document_priorities,
    "local_documents": local_documents,
    "sources": sources,
}, ensure_ascii=False), encoding="utf-8")

print(json.dumps({
    "companies": len(companies),
    "fields": len(fields),
    "document_priorities": len(document_priorities),
    "local_documents": len(local_documents),
    "sources": len(sources),
}, indent=2))
